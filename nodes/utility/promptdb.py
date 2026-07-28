# nodes/utility/promptdb.py
"""TJ_PromptDB — log prompts/settings/thumbnails to an Excel workbook and browse
them back into a workflow via a gallery-style loader node.

Two nodes:
  TJ_PromptDBSave — appends one row per image in the batch to excel_path
  TJ_PromptDBLoader — thumbnail-grid browser; outputs the selected row's fields
"""

import base64
import ipaddress
import json
import os
import re
import time
from datetime import datetime
from contextlib import contextmanager
from io import BytesIO
from typing import Any, Dict, List, Optional
from urllib.parse import quote

from PIL import Image

import folder_paths

try:
    from ...core.tj_types import any_type
except ImportError:
    # Loaded outside the package (test harnesses import this file by path), so the relative
    # import has no parent. Same tiny wildcard other nodes in this pack declare locally.
    class AnyType(str):
        def __ne__(self, _other: object) -> bool:
            return False

    any_type = AnyType("*")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = load_workbook = XLImage = Font = get_column_letter = None

try:
    from server import PromptServer
    from aiohttp import web
except Exception:
    PromptServer = None
    web = None

TJ_PROMPT_PIPE = "TJ_PROMPT_PIPE"

# Fixed column order — Logger writes it, Loader reads it. Both nodes must agree, so this
# is the single source of truth for the layout (col index is position in this list, 1-based).
# Sampler/Scheduler are appended AFTER the original 11 columns (not inserted earlier in the
# list) so workbooks created before this field existed keep reading correctly — their
# existing columns never shift position.
HEADERS = ["ID", "날짜", "Positive Prompt", "Negative Prompt", "Model", "Seed", "Steps", "CFG",
           "기타 설정", "썸네일", "메모", "Sampler", "Scheduler"]
(COL_ID, COL_DATE, COL_POS, COL_NEG, COL_MODEL, COL_SEED, COL_STEPS, COL_CFG, COL_EXTRA, COL_THUMB,
 COL_NOTE, COL_SAMPLER, COL_SCHEDULER) = range(1, 14)

# Pushed after a successful write so any open PromptDBLoader looking at the same workbook
# can refresh itself — otherwise the gallery keeps showing a stale list until the user
# remembers to hit ⟳ after every run.
UPDATED_EVENT = "tj_promptdb_updated"

THUMB_SUBDIR = "_tj_thumbnails_tmp"
SAVE_RETRIES = 3
SAVE_RETRY_DELAY = 0.6

# ── Path sandbox ───────────────────────────────────────────────────────────
# Every workbook this node reads or writes must live under <custom_node>/promptDB.
# Subfolders inside it are fine; anything that resolves outside is refused.
#
# This matters because the node's own widget value AND the loopback HTTP API both feed
# user-controlled strings straight into open()/save(). Without a jail, "../../../..." or a
# bare "C:\Windows\..." would let a workflow (or any page able to reach 127.0.0.1) read and
# overwrite arbitrary files as the ComfyUI process. Containment is checked AFTER
# os.path.realpath so symlinks/junctions pointing outside can't be used to tunnel out.
PROMPTDB_ROOT = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "promptDB",
)


# Any C0 control character, not just NUL. JSON can carry these as \b, \t, \f, …,
# so a name that looks harmless in a request body can still reach the filesystem with
# embedded control bytes — Windows rejects them at makedirs() with a confusing WinError,
# and on other platforms they'd silently create unusable names.
_CONTROL_CHARS = re.compile(r"[\x00-\x1f\x7f]")


def _ensure_root() -> str:
    os.makedirs(PROMPTDB_ROOT, exist_ok=True)
    return os.path.realpath(PROMPTDB_ROOT)


def _contain(path: str, root: str) -> bool:
    """True when `path` is `root` itself or genuinely inside it.

    Uses commonpath, not startswith: "<root>evil" starts with "<root>" as a string but is a
    sibling directory, not a child.
    """
    try:
        return os.path.commonpath([os.path.normcase(path), os.path.normcase(root)]) == os.path.normcase(root)
    except ValueError:
        return False  # different drives on Windows


def _resolve_in_root(raw: str, *, suffix: Optional[str] = None) -> str:
    """Resolve a user-supplied path inside PROMPTDB_ROOT, or raise ValueError."""
    text = str(raw or "").strip()
    if _CONTROL_CHARS.search(text):
        raise ValueError("Invalid path")
    root = _ensure_root()

    # Reject the shell-ish expansions outright rather than expanding them: %USERPROFILE% or
    # ~ can only ever point outside the sandbox, so silently rewriting them would be
    # confusing, and expanding-then-checking is easy to get subtly wrong.
    if "~" in text or "%" in text or "$" in text:
        raise ValueError("Path variables are not allowed in PromptDB paths")

    # Refuse dots-only components ("..", "....", ...). ".." is the obvious traversal, and
    # Windows strips trailing dots from names, so "...." is ambiguous enough that different
    # layers can disagree about what it means — safest to never let one through.
    for part in re.split(r"[\\/]+", text):
        if part and set(part) == {"."}:
            raise ValueError("Invalid path component")

    candidate = text if os.path.isabs(text) else os.path.join(root, text)
    resolved = os.path.realpath(candidate)
    if not _contain(resolved, root):
        raise ValueError("PromptDB paths must stay inside the promptDB folder")
    if suffix and not resolved.lower().endswith(suffix):
        raise ValueError(f"Expected a {suffix} file")
    return resolved


def _resolve_excel_path(excel_path: str) -> str:
    return _resolve_in_root(excel_path, suffix=".xlsx")


def _rel_to_root(path: str) -> str:
    """Display/storage form: path relative to the sandbox root (portable across machines)."""
    try:
        return os.path.relpath(path, _ensure_root()).replace("\\", "/")
    except ValueError:
        return path


def _thumb_dir_for(excel_path: str) -> str:
    """Thumbnail sidecar for ONE workbook: <folder>/_tj_thumbnails_tmp/<workbook name>/.

    The per-workbook subfolder is essential. Thumbnails are named row_<id>.png, and row ids
    restart at 1 in every workbook — so a single shared folder made libraries in the same
    directory overwrite each other's images, show each other's thumbnails, and delete each
    other's files on row delete. The sandbox puts every library in one folder, which makes
    that collision the normal case rather than an edge case.
    """
    stem = os.path.splitext(os.path.basename(excel_path))[0]
    return os.path.join(os.path.dirname(excel_path), THUMB_SUBDIR, stem)


def _legacy_thumb_dir(excel_path: str) -> str:
    return os.path.join(os.path.dirname(excel_path), THUMB_SUBDIR)


# Folders already swept this process. Files are only ever written to the namespaced
# location now, so nothing can reappear at the flat level while the server runs — without
# this guard, any leftover ambiguous file would make every single list_rows call re-open
# every workbook in the folder.
_MIGRATED_FOLDERS: set = set()


def _migrate_legacy_thumbnails(excel_path: str) -> None:
    """Moves pre-namespacing thumbnails into their owning workbook's subfolder.

    Only migrates a file when exactly one workbook in the folder actually contains that row
    id — if two workbooks both have a row 3, there is no way to tell whose row_3.png it is,
    so it is left alone rather than guessed at. Cheap no-op once the flat folder is empty.
    """
    legacy = _legacy_thumb_dir(excel_path)
    key = os.path.normcase(legacy)
    if key in _MIGRATED_FOLDERS:
        return
    _MIGRATED_FOLDERS.add(key)
    if not os.path.isdir(legacy):
        return
    try:
        flat = [n for n in os.listdir(legacy)
                if n.lower().endswith(".png") and os.path.isfile(os.path.join(legacy, n))]
    except OSError:
        return
    if not flat:
        return

    # Ownership is decided by which workbook has an image EMBEDDED on that row, not merely
    # a row with that id. Row ids restart at 1 everywhere, so "has row 1" is true of almost
    # every workbook and would make every file look ambiguous; an anchored picture is the
    # actual evidence that this workbook is the one the thumbnail was written for.
    folder = os.path.dirname(excel_path)
    owners: Dict[str, set] = {}
    for name in os.listdir(folder):
        if not name.lower().endswith(".xlsx"):
            continue
        book = os.path.join(folder, name)
        try:
            wb = load_workbook(book)  # not read_only: that mode does not load images
            ws = wb.active
            id_by_row = {r[0].row: r[COL_ID - 1].value for r in ws.iter_rows(min_row=2)}
            ids = set()
            for img in getattr(ws, "_images", []):
                row_no = img.anchor._from.row + 1
                value = id_by_row.get(row_no)
                if isinstance(value, (int, float)):
                    ids.add(int(value))
            wb.close()
        except Exception:
            continue
        if ids:
            owners[book] = ids

    for filename in flat:
        match = re.fullmatch(r"row_(\d+)\.png", filename)
        if not match:
            continue
        row_id = int(match.group(1))
        claimants = [book for book, ids in owners.items() if row_id in ids]
        if len(claimants) != 1:
            continue  # nobody owns it, or it is ambiguous — leave it untouched
        target_dir = _thumb_dir_for(claimants[0])
        os.makedirs(target_dir, exist_ok=True)
        target = os.path.join(target_dir, filename)
        if os.path.exists(target):
            continue
        try:
            os.replace(os.path.join(legacy, filename), target)
        except OSError:
            pass


def _recover_missing_thumbnails(excel_path: str) -> int:
    """Rebuilds absent sidecar thumbnails from the images embedded in the workbook.

    The workbook is the durable copy — openpyxl embeds each thumbnail into the sheet — while
    the sidecar folder is just a cache the gallery reads from. They can drift apart: the
    folder gets cleaned up, the .xlsx is copied somewhere without it, or a backup restores
    only the workbook. Re-extracting means the gallery recovers instead of showing blanks.
    Returns how many files were written.
    """
    thumb_dir = _thumb_dir_for(excel_path)
    try:
        wb = load_workbook(excel_path)  # read_only mode does not load images
    except Exception:
        return 0
    recovered = 0
    try:
        ws = wb.active
        id_by_row = {row[0].row: row[COL_ID - 1].value for row in ws.iter_rows(min_row=2)}
        for img in getattr(ws, "_images", []):
            value = id_by_row.get(img.anchor._from.row + 1)
            if not isinstance(value, (int, float)):
                continue
            target = os.path.join(thumb_dir, f"row_{int(value)}.png")
            if os.path.exists(target):
                continue
            try:
                data = img._data()
                os.makedirs(thumb_dir, exist_ok=True)
                # Normalise through PIL: the embedded blob may be any format openpyxl
                # accepted, but the sidecar is always read back as PNG.
                with Image.open(BytesIO(data)) as im:
                    im.convert("RGB").save(target, format="PNG")
                recovered += 1
            except Exception:
                continue
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return recovered


# Header labels that were renamed after workbooks already existed. Only the label changed —
# same column, same data — so these are relabelled in place rather than migrated.
_RENAMED_HEADERS = {COL_NOTE: {"원본 경로"}}


def _sync_headers(ws) -> None:
    """Brings an existing sheet's header row up to date with HEADERS.

    Runs on every write path, not just when the Save node appends: someone who only ever
    edits rows in the gallery would otherwise keep seeing the old column label in Excel.
    Row data is never touched.
    """
    for col in (COL_SAMPLER, COL_SCHEDULER, COL_NOTE):
        current = ws.cell(row=1, column=col).value
        if current is None or current in _RENAMED_HEADERS.get(col, set()):
            ws.cell(row=1, column=col, value=HEADERS[col - 1]).font = Font(bold=True)


def _open_or_create_workbook(excel_path: str):
    if os.path.isfile(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
        _sync_headers(ws)
        return wb, ws
    os.makedirs(os.path.dirname(excel_path) or ".", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Prompts"
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    widths = {COL_ID: 6, COL_DATE: 12, COL_POS: 42, COL_NEG: 32, COL_MODEL: 18,
              COL_SEED: 10, COL_STEPS: 8, COL_CFG: 8, COL_EXTRA: 24, COL_THUMB: 10, COL_NOTE: 32,
              COL_SAMPLER: 16, COL_SCHEDULER: 14}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    return wb, ws


def _next_id(ws) -> int:
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        val = row[0]
        if isinstance(val, (int, float)):
            max_id = max(max_id, int(val))
    return max_id + 1


def _save_with_retry(wb, excel_path: str) -> None:
    last_exc = None
    for attempt in range(SAVE_RETRIES):
        try:
            wb.save(excel_path)
            return
        except PermissionError as exc:
            last_exc = exc
            time.sleep(SAVE_RETRY_DELAY)
        except Exception as exc:
            last_exc = exc
            time.sleep(SAVE_RETRY_DELAY)
    raise RuntimeError(
        f"Could not save {excel_path} after {SAVE_RETRIES} attempts "
        f"(is it open in Excel or another program?): {last_exc}"
    )


@contextmanager
def _open_readonly(path: str):
    """Opens a workbook read-only and ALWAYS closes it.

    openpyxl's read_only mode streams from the file and keeps the handle open until
    close() is called. On Windows that handle is a lock: after the gallery listed a
    library, saving to that same workbook would fail with a PermissionError that looks
    exactly like "the file is open in Excel", and the file couldn't be deleted either.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        yield wb
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _notify_updated(excel_path: str) -> None:
    """Tell any open Loader that this workbook changed. Best-effort — never fails a save."""
    try:
        sender = getattr(getattr(PromptServer, "instance", None), "send_sync", None)
        if sender:
            sender(UPDATED_EVENT, {"path": _rel_to_root(excel_path)})
    except Exception:
        pass


def _tensor_to_pil(image_slice: Any) -> Image.Image:
    img = image_slice[0] if hasattr(image_slice, "dim") and image_slice.dim() == 4 else image_slice
    arr = (img.detach().cpu().numpy() * 255.0).clip(0, 255).astype("uint8")
    return Image.fromarray(arr)


# ── Auto-extract: walk the executing prompt graph to find the settings that
# actually produced this image ─────────────────────────────────────────────
#
# ComfyUI hands each node the whole executing prompt (PROMPT) plus its own node id
# (UNIQUE_ID). In that structure a linked input is ["<origin_node_id>", <slot>], so the
# graph can be walked upstream from our own `images` input. That beats asking the user to
# pick a sampler from a dropdown: whichever sampler produced this image is already encoded
# in the links, it stays correct when node ids change, and it works for API runs too.

# Detection is by input SHAPE, not class name, so third-party samplers/loaders are caught
# as well — matching on names like "KSampler" would only cover core nodes.
_SAMPLER_MARKERS = ({"seed", "steps", "cfg"}, {"sampler_name", "scheduler"})
_LATENT_KEYS = ("latent_image", "samples", "latent")


def _is_sampler(inputs: Dict[str, Any]) -> bool:
    """A sampler both *looks* like one and actually consumes a latent.

    The marker widgets alone are not enough: all-in-one loaders and settings hubs
    (e.g. "Smart Model Loader [Eclipse]") expose seed/steps/cfg/sampler_name/scheduler as
    plain widgets without sampling anything, and matching those made such a node get
    reported as a refiner pass. Taking a latent in is what actually distinguishes a
    sampler, and every KSampler/SamplerCustom variant has one.
    """
    keys = set(inputs.keys())
    if not any(marker <= keys for marker in _SAMPLER_MARKERS):
        return False
    return any(k in keys for k in _LATENT_KEYS)


def _link_origin(value: Any) -> Optional[str]:
    """A linked input is [origin_node_id, output_slot]; a literal widget value is not."""
    if isinstance(value, list) and len(value) == 2 and isinstance(value[0], (str, int)):
        return str(value[0])
    return None


def _collect_sampler_chain(prompt: Dict[str, Any], start_id: str) -> List[str]:
    """Every sampler upstream of start_id, ordered downstream-last (main .. refiners).

    Walks breadth-first through *all* upstream links, not just latent ones, so it still
    finds the chain when a VAEDecode/upscale/etc. sits between us and the sampler.
    """
    chain: List[str] = []
    seen = set()
    frontier = [start_id]
    while frontier:
        node_id = frontier.pop(0)
        if node_id in seen or node_id not in prompt:
            continue
        seen.add(node_id)
        inputs = prompt[node_id].get("inputs") or {}
        if _is_sampler(inputs) and node_id != start_id:
            chain.append(node_id)
        for value in inputs.values():
            origin = _link_origin(value)
            if origin is not None:
                frontier.append(origin)
    # Frontier order is downstream-first (nearest sampler = the last refiner), so reverse
    # to get execution order: main sampler first, refinement passes after it.
    chain.reverse()
    return chain


def _starts_from_empty_latent(prompt: Dict[str, Any], sampler_id: str) -> bool:
    """True when this sampler denoises from scratch rather than refining an existing latent.

    That is what distinguishes the MAIN sampler from a refiner pass: the main one starts at
    an EmptyLatentImage (denoise 1.0), refiners take another sampler's latent at denoise<1.
    Recording a refiner's seed/steps/cfg in the main columns would make rows incomparable —
    some rows would hold "10 steps, cfg 4.0, denoise 0.35" and others the real settings.
    """
    inputs = prompt.get(sampler_id, {}).get("inputs") or {}
    try:
        if float(inputs.get("denoise", 1.0)) < 1.0:
            return False
    except (TypeError, ValueError):
        pass
    for key in _LATENT_KEYS:
        origin = _link_origin(inputs.get(key))
        if origin is None:
            continue
        origin_class = str(prompt.get(origin, {}).get("class_type") or "")
        if "EmptyLatent" in origin_class or "EmptySD3Latent" in origin_class:
            return True
        # Anything that is itself a sampler means we are the refinement pass.
        if _is_sampler(prompt.get(origin, {}).get("inputs") or {}):
            return False
    return True


def _find_upstream(prompt: Dict[str, Any], start_id: str, predicate) -> Optional[str]:
    seen = set()
    frontier = [start_id]
    while frontier:
        node_id = frontier.pop(0)
        if node_id in seen or node_id not in prompt:
            continue
        seen.add(node_id)
        node = prompt[node_id]
        if node_id != start_id and predicate(node):
            return node_id
        for value in (node.get("inputs") or {}).values():
            origin = _link_origin(value)
            if origin is not None:
                frontier.append(origin)
    return None


# ComfyUI represents "this slot is unused" as the literal string "None" in combo widgets,
# so a loader that runs off unet_name still carries ckpt_name="None". Taking the first
# non-empty key would record the placeholder as if it were the model.
_PLACEHOLDER_NAMES = {"", "none", "null", "undefined"}
_MODEL_NAME_KEYS = ("ckpt_name", "unet_name", "model_name", "model", "base_model")


def _model_name_of(inputs: Dict[str, Any]) -> str:
    """The real model file named by a loader node, ignoring unset placeholders."""
    for key in _MODEL_NAME_KEYS:
        value = inputs.get(key)
        if _link_origin(value) is not None or not isinstance(value, str):
            continue
        if value.strip().lower() in _PLACEHOLDER_NAMES:
            continue
        return value
    return ""


_RESOLVE_MAX_DEPTH = 8


def _resolve_setting(prompt: Dict[str, Any], node_id: Optional[str], key: str,
                     depth: int = 0, seen: Optional[set] = None):
    """The value of `key` for a node, following links upstream when it isn't a literal.

    Plenty of workflows don't type settings on the sampler itself: an all-in-one loader or
    a "pipe"/hub node holds seed/steps/cfg and feeds them in, so reading only the sampler's
    own widgets finds nothing but links. Walking upstream for a widget of the SAME name
    recovers the value the run actually used. Bounded by depth and a visited set — the
    search only ever moves toward inputs, so it stays inside the sampler's own ancestry.
    """
    if node_id is None or depth > _RESOLVE_MAX_DEPTH or node_id not in prompt:
        return None
    if seen is None:
        seen = set()
    if node_id in seen:
        return None
    seen.add(node_id)

    inputs = prompt[node_id].get("inputs") or {}
    if key in inputs:
        value = inputs[key]
        origin = _link_origin(value)
        if origin is None:
            return value
        found = _resolve_setting(prompt, origin, key, depth + 1, seen)
        if found is not None:
            return found

    for value in inputs.values():
        origin = _link_origin(value)
        if origin is None:
            continue
        found = _resolve_setting(prompt, origin, key, depth + 1, seen)
        if found is not None:
            return found
    return None


def _literal(prompt: Dict[str, Any], node_id: Optional[str], key: str, default=None):
    """Reads a widget value, ignoring it when the input is a link rather than a literal."""
    if not node_id:
        return default
    value = (prompt.get(node_id, {}).get("inputs") or {}).get(key, default)
    return default if _link_origin(value) is not None else value


def _extract_from_prompt(prompt: Dict[str, Any], self_id: str) -> Dict[str, Any]:
    """Best-effort settings for the image reaching this node. Never raises."""
    result: Dict[str, Any] = {}
    chain = _collect_sampler_chain(prompt, self_id)
    if not chain:
        return result

    # Prefer the sampler that denoises from scratch; if none does (img2img and friends),
    # fall back to the earliest one in execution order.
    main_id = next((sid for sid in chain if _starts_from_empty_latent(prompt, sid)), chain[0])
    seed = _resolve_setting(prompt, main_id, "seed")
    result["seed"] = seed if seed is not None else _resolve_setting(prompt, main_id, "noise_seed")
    result["steps"] = _resolve_setting(prompt, main_id, "steps")
    result["cfg"] = _resolve_setting(prompt, main_id, "cfg")
    result["sampler_name"] = _resolve_setting(prompt, main_id, "sampler_name")
    result["scheduler"] = _resolve_setting(prompt, main_id, "scheduler")

    refiners = [sid for sid in chain if sid != main_id]
    if refiners:
        parts = []
        for sid in refiners:
            cls = prompt.get(sid, {}).get("class_type") or "Sampler"
            bits = [f"{key}={_literal(prompt, sid, key)}"
                    for key in ("steps", "cfg", "denoise", "sampler_name")
                    if _literal(prompt, sid, key) is not None]
            parts.append(f"refiner: {cls} " + " ".join(bits) if bits else f"refiner: {cls}")
        result["extra_settings"] = "\n".join(parts)

    ckpt_id = _find_upstream(prompt, self_id, lambda n: bool(_model_name_of(n.get("inputs") or {})))
    if ckpt_id:
        result["model_name"] = _model_name_of(prompt[ckpt_id].get("inputs") or {})

    return {k: v for k, v in result.items() if v is not None and v != ""}


class TJ_PromptDBSave:
    DESCRIPTION = "Appends one row per image (prompt/settings/thumbnail) to an Excel workbook."

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "images": ("IMAGE",),
                "positive_prompt": ("STRING", {"multiline": True, "default": ""}),
                "excel_path": ("STRING", {"default": ""}),
                "thumbnail_size": ("INT", {"default": 128, "min": 16, "max": 512, "step": 8}),
                # BYPASS lets the node stay wired in the graph while writing nothing —
                # images pass straight through, so downstream nodes are unaffected.
                "mode": ("BOOLEAN", {"default": True, "label_on": "SAVE", "label_off": "BYPASS"}),
                "auto_extract": ("BOOLEAN", {"default": True, "label_on": "auto", "label_off": "manual"}),
            },
            "optional": {
                "negative_prompt": ("STRING", {"multiline": True, "default": ""}),
                "model_name": ("STRING", {"default": ""}),
                "seed": ("INT", {"default": 0, "min": 0, "max": 0xffffffffffffffff}),
                # Defaults double as the manual-mode fallback: when auto-extract is off, or
                # is on but can't find a value in the graph, these are what get recorded —
                # a blank/0 row would be worse than a sane default.
                "steps": ("INT", {"default": 8, "min": 0, "max": 10000}),
                "cfg": ("FLOAT", {"default": 1.0, "min": 0.0, "max": 100.0, "step": 0.1}),
                "sampler_name": ("STRING", {"default": "euler"}),
                "scheduler": ("STRING", {"default": "simple"}),
                "extra_settings": ("STRING", {"multiline": True, "default": ""}),
                # Free-form note. Was "source_path" (an image path nothing ever filled
                # in); it is now whatever the user wants to carry alongside the row.
                "note": ("STRING", {"multiline": True, "default": ""}),
                # TJ_NODE's embedded Set/Get — handled entirely by web/set_getnode_tj.js,
                # which keys off these exact widget names:
                #   get_name      — receive `images` wirelessly from a named provider
                #   setnode_name  — publish this node's own output under a name
                # Appended LAST on purpose: ComfyUI restores widget values by POSITION, so
                # inserting a widget anywhere else shifts every later value in workflows
                # saved before it existed (seen live: positive_prompt=128, steps="euler").
                "get_name": (["(none)"], {"default": "(none)"}),
                "setnode_name": ("STRING", {"default": "PromptDB"}),
            },
            "hidden": {"prompt": "PROMPT", "unique_id": "UNIQUE_ID"},
        }

    RETURN_TYPES = ("IMAGE",)
    RETURN_NAMES = ("images",)
    FUNCTION = "log"
    CATEGORY = " ✨ TJ_Node/Utility"
    OUTPUT_NODE = True

    @classmethod
    def VALIDATE_INPUTS(cls, **kwargs):
        # get_name is declared as ["(none)"] but the front end fills it with whatever Set
        # names exist in the graph, so server-side combo validation would reject a perfectly
        # valid selection. Same bypass every other embedded-Get node in this pack uses.
        return True

    # These defaults MUST match INPUT_TYPES. When an optional input is left unconnected
    # ComfyUI omits the kwarg entirely, so it is the signature default that gets recorded —
    # not the widget default the user sees on the node. Having them disagree wrote 0/"" into
    # the sheet for a disconnected socket instead of the promised 8 / 1.0 / euler / simple.
    def log(self, images, positive_prompt, excel_path, thumbnail_size, mode=True, auto_extract=True,
            negative_prompt="", model_name="", seed=0, steps=8, cfg=1.0,
            sampler_name="euler", scheduler="simple", extra_settings="", note="",
            get_name="(none)", setnode_name="PromptDB", prompt=None, unique_id=None):
        # BYPASS: hand the images straight back, touching nothing. Deliberately checked
        # before any validation so an unconfigured excel_path can't fail a bypassed run.
        if not mode:
            return (images,)

        if Workbook is None:
            raise RuntimeError("openpyxl is required for TJ_PromptDBSave. pip install openpyxl")

        path = _resolve_excel_path(excel_path)
        if not path:
            raise ValueError("excel_path is required.")

        if auto_extract and prompt and unique_id is not None:
            try:
                found = _extract_from_prompt(prompt, str(unique_id))
            except Exception:
                found = {}  # auto-extract is a convenience; never fail the save over it

            # Precedence in auto mode: a CONNECTED input always wins (wiring something up is
            # unambiguously deliberate), otherwise the extracted value wins, otherwise the
            # widget's own value stands — which for steps/cfg/sampler/scheduler is a usable
            # default rather than 0/"" , so an un-extractable run still records sane settings.
            # Widget literals can't be treated as overrides here: their defaults are now
            # non-empty, so "user typed 8" and "nobody touched it" look identical.
            linked = {k for k, v in (prompt.get(str(unique_id), {}).get("inputs") or {}).items()
                      if _link_origin(v) is not None}

            def pick(key, current):
                if key in linked or key not in found:
                    return current
                return found[key]

            model_name = pick("model_name", model_name)
            seed = pick("seed", seed)
            steps = pick("steps", steps)
            cfg = pick("cfg", cfg)
            sampler_name = pick("sampler_name", sampler_name)
            scheduler = pick("scheduler", scheduler)
            # extra_settings is a free-form notes field rather than a comparable value, so
            # refiner details are appended instead of being suppressed by the user's text.
            refiner_note = found.get("extra_settings", "")
            if refiner_note:
                extra_settings = f"{extra_settings}\n{refiner_note}".strip() if extra_settings else refiner_note

        wb, ws = _open_or_create_workbook(path)
        try:
            _migrate_legacy_thumbnails(path)
        except Exception:
            pass  # best-effort; a failed migration must not stop the row being logged
        thumb_dir = _thumb_dir_for(path)
        os.makedirs(thumb_dir, exist_ok=True)

        count = images.shape[0] if hasattr(images, "shape") and len(images.shape) == 4 else 1
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        thumb_px = int(thumbnail_size)

        for i in range(count):
            row_id = _next_id(ws)
            row_idx = ws.max_row + 1

            pil = _tensor_to_pil(images[i:i + 1] if count > 1 else images)
            pil.thumbnail((thumb_px, thumb_px), Image.Resampling.LANCZOS)
            thumb_path = os.path.join(thumb_dir, f"row_{row_id}.png")
            pil.save(thumb_path, format="PNG")

            ws.cell(row=row_idx, column=COL_ID, value=row_id)
            ws.cell(row=row_idx, column=COL_DATE, value=now_str)
            ws.cell(row=row_idx, column=COL_POS, value=str(positive_prompt or ""))
            ws.cell(row=row_idx, column=COL_NEG, value=str(negative_prompt or ""))
            ws.cell(row=row_idx, column=COL_MODEL, value=str(model_name or ""))
            ws.cell(row=row_idx, column=COL_SEED, value=int(seed))
            ws.cell(row=row_idx, column=COL_STEPS, value=int(steps))
            ws.cell(row=row_idx, column=COL_CFG, value=float(cfg))
            ws.cell(row=row_idx, column=COL_EXTRA, value=str(extra_settings or ""))
            ws.cell(row=row_idx, column=COL_NOTE, value=str(note or ""))
            ws.cell(row=row_idx, column=COL_SAMPLER, value=str(sampler_name or ""))
            ws.cell(row=row_idx, column=COL_SCHEDULER, value=str(scheduler or ""))

            ws.row_dimensions[row_idx].height = max(15, thumb_px * 0.75)
            try:
                xl_img = XLImage(thumb_path)
                xl_img.width, xl_img.height = pil.width, pil.height
                ws.add_image(xl_img, f"{get_column_letter(COL_THUMB)}{row_idx}")
            except Exception:
                pass  # thumbnail embed is best-effort; the row's data is what matters

        _save_with_retry(wb, path)
        _notify_updated(path)
        return (images,)


class TJ_PromptDBLoader:
    DESCRIPTION = "Gallery browser for a TJ_PromptDBSave workbook — click a thumbnail to load its row."

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": ""}),
            },
            "optional": {
                # JS-managed selection state, serialized with the workflow so the selection
                # survives reload. Not meant to be hand-edited.
                "selected_id": ("INT", {"default": -1, "min": -1, "max": 0x7fffffff}),
                # Auto Set — registers every output as a wireless provider so the loaded
                # row's fields can be picked up by name instead of dragging 11 wires. The
                # node type must ALSO be listed in AUTO_SET_PROVIDER_TYPES in
                # web/set_getnode_tj.js; web/promptdb_tj.js assigns the per-slot names.
                # Appended last for the same positional-restore reason as PromptDBSave.
                "auto_set": ("BOOLEAN", {"default": False,
                                         "label_on": "Auto Set ON", "label_off": "Auto Set OFF"}),
            },
        }

    # Deliberately just two sockets. Everything else comes out of PromptDBBridge(TJ), which
    # you place next to whatever consumes the values — one wire crosses the canvas instead
    # of ten.
    #
    # An earlier version exposed all eleven and let the UI collapse them. That was broken:
    # hiding an output removes it from the frontend's `outputs` array, which renumbers the
    # slots that get serialised into the prompt, while the backend still resolves those
    # indices against RETURN_TYPES. A link from the collapsed "pipe" (visually slot 1) was
    # sent as slot 1 and read server-side as negative_prompt — "Return type mismatch:
    # received_type(STRING) mismatch input_type(TJ_PROMPT_PIPE)". Keeping the real socket
    # list short and constant removes that whole class of bug.
    RETURN_TYPES = ("STRING", TJ_PROMPT_PIPE)
    RETURN_NAMES = ("positive_prompt", "pipe")
    FUNCTION = "load"
    CATEGORY = " ✨ TJ_Node/Utility"

    def load(self, excel_path, selected_id=-1, auto_set=False):
        if load_workbook is None:
            raise RuntimeError("openpyxl is required for TJ_PromptDBLoader. pip install openpyxl")

        path = _resolve_excel_path(excel_path)
        if not path or not os.path.isfile(path):
            raise FileNotFoundError(f"Excel file not found: {path}")
        if int(selected_id) < 0:
            raise RuntimeError("No row selected — click a thumbnail in the TJ_PromptDBLoader grid first.")

        with _open_readonly(path) as wb:
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[COL_ID - 1] == int(selected_id):
                    positive = str(row[COL_POS - 1] or "")
                    negative = str(row[COL_NEG - 1] or "")
                    model = str(row[COL_MODEL - 1] or "")
                    seed = int(row[COL_SEED - 1] or 0)
                    steps = int(row[COL_STEPS - 1] or 0)
                    cfg = float(row[COL_CFG - 1] or 0.0)
                    extra = str(row[COL_EXTRA - 1] or "")
                    note_text = str(row[COL_NOTE - 1] or "")
                    sampler_name = str(row[COL_SAMPLER - 1] or "") if len(row) >= COL_SAMPLER else ""
                    scheduler = str(row[COL_SCHEDULER - 1] or "") if len(row) >= COL_SCHEDULER else ""
                    pipe = {
                        "positive_prompt": positive, "negative_prompt": negative, "model_name": model,
                        "seed": seed, "steps": steps, "cfg": cfg, "sampler_name": sampler_name, "scheduler": scheduler,
                        "extra_settings": extra, "note": note_text,
                    }
                    return (positive, pipe)

        raise RuntimeError(f"Row ID {selected_id} not found in {path} (was it deleted from the sheet?).")


class TJ_PromptDBBridge:
    """Fans a TJ_PROMPT_PIPE back out into individual fields.

    Lets the Loader stay compact — it only needs to expose positive_prompt and pipe — while
    this node, placed next to whatever consumes the values, unpacks the rest. One wire
    crosses the canvas instead of ten, and with Auto Set on it republishes every field
    wirelessly from wherever it sits.
    """

    DESCRIPTION = "Unpacks a PromptDB pipe into its individual fields."

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "pipe": (TJ_PROMPT_PIPE,),
            },
            "optional": {
                "get_name": (["(none)"], {"default": "(none)"}),
                "auto_set": ("BOOLEAN", {"default": False,
                                         "label_on": "Auto Set ON", "label_off": "Auto Set OFF"}),
            },
        }

    # sampler_name / scheduler / model_name are wildcards, not STRING. Their real consumers
    # (KSampler.sampler_name, KSampler.scheduler, a checkpoint loader's ckpt_name) are COMBO
    # inputs, and LiteGraph refuses STRING -> COMBO, so typing them as STRING made exactly
    # the connections this node exists for impossible. AnyType also satisfies the backend's
    # `received_type != input_type` check, and still connects to plain STRING consumers.
    # `pipe` is passed through last so bridges can be chained.
    RETURN_TYPES = ("STRING", "STRING", any_type, "INT", "INT", "FLOAT",
                    "STRING", any_type, any_type, any_type, TJ_PROMPT_PIPE)
    RETURN_NAMES = ("positive_prompt", "negative_prompt", "model_name", "seed", "steps", "cfg",
                    "extra_settings", "note", "sampler_name", "scheduler", "pipe")
    FUNCTION = "unpack"
    CATEGORY = " ✨ TJ_Node/Utility"

    @classmethod
    def VALIDATE_INPUTS(cls, **kwargs):
        return True

    def unpack(self, pipe, get_name="(none)", auto_set=False):
        data = pipe if isinstance(pipe, dict) else {}

        def text(key):
            return str(data.get(key) or "")

        def number(key, cast, default):
            try:
                return cast(data.get(key, default) or default)
            except (TypeError, ValueError):
                return default

        return (
            text("positive_prompt"), text("negative_prompt"), text("model_name"),
            number("seed", int, 0), number("steps", int, 0), number("cfg", float, 0.0),
            text("extra_settings"), text("note"),
            text("sampler_name"), text("scheduler"),
            data,
        )


# ── Local API: gallery listing + row update for the Loader's JS grid ───────────

def _is_loopback(host):
    if not host:
        return False
    try:
        ip = ipaddress.ip_address(host)
    except ValueError:
        return False
    mapped = getattr(ip, "ipv4_mapped", None)
    if mapped is not None:
        ip = mapped
    return ip.is_loopback


def _local_only(request):
    """Block non-loopback and cross-origin requests (same guard as Shortcut Launcher's).

    These routes read, write, copy and create files on disk. Two distinct threats:
    a remote caller reaching the port, and a malicious page open in the user's own browser
    silently POSTing to 127.0.0.1 (CSRF) — the latter *passes* a loopback check, since the
    request really does originate locally. Requiring a same-origin Origin/Referer closes it
    while the node's own UI, served from this very origin, keeps working.

    The promptDB sandbox already caps the blast radius to that one folder; this keeps a
    drive-by page from trashing the user's libraries inside it.
    """
    if not _is_loopback(request.remote):
        return web.json_response({"ok": False, "error": "Local (loopback) requests only."}, status=403)

    host_header = request.headers.get("Host", "")
    origin = request.headers.get("Origin") or request.headers.get("Referer") or ""
    if host_header and origin:
        origin_host = origin.split("://", 1)[-1].split("/", 1)[0]
        if origin_host != host_header:
            return web.json_response({"ok": False, "error": "Cross-origin requests are not allowed."}, status=403)
    return None


def _thumbnail_data_uri(thumb_path: str) -> str:
    try:
        with open(thumb_path, "rb") as f:
            return "data:image/png;base64," + base64.b64encode(f.read()).decode("ascii")
    except Exception:
        return ""


async def _handle_list_rows(request):
    blocked = _local_only(request)
    if blocked is not None:
        return blocked
    try:
        payload = await request.json()
        path = _resolve_excel_path(payload.get("excel_path", ""))
        if not path or not os.path.isfile(path):
            return web.json_response({"rows": []})

        try:
            _migrate_legacy_thumbnails(path)
        except Exception:
            pass  # best-effort self-healing; never block listing over it

        def read_rows():
            thumb_dir = _thumb_dir_for(path)
            out = []
            with _open_readonly(path) as wb:
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or row[COL_ID - 1] is None:
                        continue
                    row_id = int(row[COL_ID - 1])
                    out.append({
                        "id": row_id,
                        "date": str(row[COL_DATE - 1] or ""),
                        "positive_prompt": str(row[COL_POS - 1] or ""),
                        "negative_prompt": str(row[COL_NEG - 1] or ""),
                        "model_name": str(row[COL_MODEL - 1] or ""),
                        "seed": row[COL_SEED - 1],
                        "steps": row[COL_STEPS - 1],
                        "cfg": row[COL_CFG - 1],
                        "extra_settings": str(row[COL_EXTRA - 1] or ""),
                        "note": str(row[COL_NOTE - 1] or ""),
                        "sampler_name": str(row[COL_SAMPLER - 1] or "") if len(row) >= COL_SAMPLER else "",
                        "scheduler": str(row[COL_SCHEDULER - 1] or "") if len(row) >= COL_SCHEDULER else "",
                        "thumbnail": _thumbnail_data_uri(os.path.join(thumb_dir, f"row_{row_id}.png")),
                    })
            return out

        rows = read_rows()
        # Recovery loads the workbook again in full (read_only mode can't see images), so
        # only pay for it when a thumbnail is actually missing — and re-read only if it
        # managed to restore something.
        if any(not r["thumbnail"] for r in rows):
            try:
                if _recover_missing_thumbnails(path):
                    rows = read_rows()
            except Exception:
                pass

        rows.sort(key=lambda r: r["id"], reverse=True)
        return web.json_response({"rows": rows})
    except Exception as exc:
        return web.json_response({"rows": [], "error": str(exc)}, status=500)


async def _handle_update_row(request):
    blocked = _local_only(request)
    if blocked is not None:
        return blocked
    try:
        payload = await request.json()
        path = _resolve_excel_path(payload.get("excel_path", ""))
        row_id = int(payload.get("id", -1))
        fields = payload.get("fields") or {}
        if not path or not os.path.isfile(path):
            return web.json_response({"ok": False, "error": "Excel file not found"}, status=404)

        wb = load_workbook(path)
        ws = wb.active
        target_row = None
        for row in ws.iter_rows(min_row=2):
            cell = row[COL_ID - 1]
            if cell.value == row_id:
                target_row = row[0].row
                break
        if target_row is None:
            return web.json_response({"ok": False, "error": f"Row ID {row_id} not found"}, status=404)

        field_col = {
            "positive_prompt": COL_POS, "negative_prompt": COL_NEG, "model_name": COL_MODEL,
            "seed": COL_SEED, "steps": COL_STEPS, "cfg": COL_CFG,
            "extra_settings": COL_EXTRA, "note": COL_NOTE,
            "sampler_name": COL_SAMPLER, "scheduler": COL_SCHEDULER,
        }
        for key, col in field_col.items():
            if key not in fields:
                continue
            value = fields[key]
            if key == "seed":
                value = int(value or 0)
            elif key == "steps":
                value = int(value or 0)
            elif key == "cfg":
                value = float(value or 0.0)
            else:
                value = str(value or "")
            ws.cell(row=target_row, column=col, value=value)

        _sync_headers(ws)
        _save_with_retry(wb, path)
        _notify_updated(path)
        return web.json_response({"ok": True})
    except Exception as exc:
        return web.json_response({"ok": False, "error": str(exc)}, status=400)


# ── Library management: named shortcuts to excel_path files ────────────────
# Solves "have to retype/remember the full path every time" — the user picks a library
# by name from a list instead. The list itself is just {id, name, path} entries the user
# built up via "Import library" (native file picker); it's not a second data store.

def _libraries_file_path() -> str:
    return os.path.join(folder_paths.get_user_directory(), "tj_node", "promptdb_libraries.json")


def _load_libraries() -> List[Dict[str, Any]]:
    path = _libraries_file_path()
    if not os.path.isfile(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _save_libraries(libraries: List[Dict[str, Any]]) -> None:
    """Persists the library list, dropping any entry that points outside the sandbox.

    Validating on write as well as on read means a stale config (or a hand-edited one)
    can never reintroduce an out-of-sandbox path.
    """
    path = _libraries_file_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    clean = []
    for lib in libraries:
        try:
            resolved = _resolve_excel_path(lib.get("path", ""))
        except ValueError:
            continue
        clean.append({
            "id": lib.get("id"),
            "name": lib.get("name"),
            # Stored relative to the sandbox root so the config stays portable.
            "path": _rel_to_root(resolved),
            "active": lib.get("active", True),
        })
    with open(path, "w", encoding="utf-8") as f:
        json.dump(clean, f, ensure_ascii=False, indent=2)


def _count_rows(path: str) -> int:
    if not path or not os.path.isfile(path):
        return 0
    try:
        with _open_readonly(path) as wb:
            ws = wb.active
            return sum(1 for row in ws.iter_rows(min_row=2, max_col=1, values_only=True)
                       if row and row[0] is not None)
    except Exception:
        return 0


async def _handle_get_libraries(request):
    blocked = _local_only(request)
    if blocked is not None:
        return blocked
    libraries = []
    for lib in _load_libraries():
        try:
            resolved = _resolve_excel_path(lib.get("path", ""))
        except ValueError:
            continue  # stale entry from before the sandbox existed — hide it
        lib["path"] = _rel_to_root(resolved)
        lib["count"] = _count_rows(resolved)
        libraries.append(lib)
    return web.json_response({"libraries": libraries, "root": _ensure_root()})


async def _handle_save_libraries(request):
    blocked = _local_only(request)
    if blocked is not None:
        return blocked
    try:
        payload = await request.json()
        libraries = payload.get("libraries")
        if not isinstance(libraries, list):
            raise TypeError("libraries must be a list")
        _save_libraries(libraries)
        return web.json_response({"ok": True})
    except Exception as exc:
        return web.json_response({"ok": False, "error": str(exc)}, status=400)


async def _handle_browse_list(request):
    """Lists subfolders and .xlsx files under PROMPTDB_ROOT for the in-app file picker.

    Replaces a native OS file dialog (the embedded Python ComfyUI ships with has no
    tkinter/Tcl, so filedialog.* raises ModuleNotFoundError). Navigation is confined to the
    sandbox: there are no drive letters, and `parent` stops at the root so the UI cannot
    offer a way out even if the caller asks for one.
    """
    blocked = _local_only(request)
    if blocked is not None:
        return blocked
    try:
        root = _ensure_root()
        raw_path = str(payload_path(await request.json()))
        try:
            path = _resolve_in_root(raw_path) if raw_path else root
        except ValueError:
            path = root  # out-of-sandbox request: silently snap back to the root
        if not os.path.isdir(path):
            path = root

        dirs, files = [], []
        try:
            with os.scandir(path) as it:
                for entry in it:
                    try:
                        if entry.name == THUMB_SUBDIR:
                            continue  # internal sidecar, not user content
                        if entry.is_dir():
                            dirs.append(entry.name)
                        elif entry.is_file() and entry.name.lower().endswith(".xlsx"):
                            files.append(entry.name)
                    except OSError:
                        continue
        except PermissionError:
            pass
        dirs.sort(key=str.lower)
        files.sort(key=str.lower)

        parent = None if os.path.normcase(path) == os.path.normcase(root) else os.path.dirname(path)
        return web.json_response({
            "ok": True,
            "path": path,
            "rel": "" if parent is None else _rel_to_root(path),
            "parent": parent,
            "root": root,
            "dirs": dirs,
            "files": files,
        })
    except Exception as exc:
        return web.json_response({"ok": False, "error": str(exc)}, status=500)


def payload_path(payload) -> str:
    return str((payload or {}).get("path") or "").strip()


# Subfolders are explicitly allowed inside the sandbox — this is how the user organises
# libraries. The name is validated as a single path segment so it can't smuggle in "..".
async def _handle_browse_mkdir(request):
    blocked = _local_only(request)
    if blocked is not None:
        return blocked
    try:
        payload = await request.json()
        parent = _resolve_in_root(payload_path(payload) or "")
        name = str(payload.get("name") or "").strip()
        if (not name or name in (".", "..")
                or re.search(r'[\\/:*?"<>|]', name)
                or _CONTROL_CHARS.search(name)
                or name.endswith(".")):  # Windows strips trailing dots — same ambiguity as "...."
            return web.json_response({"ok": False, "error": "Invalid folder name"}, status=400)

        target = _resolve_in_root(os.path.join(parent, name))
        os.makedirs(target, exist_ok=True)
        return web.json_response({"ok": True, "path": target})
    except ValueError as exc:
        return web.json_response({"ok": False, "error": str(exc)}, status=400)
    except Exception as exc:
        return web.json_response({"ok": False, "error": str(exc)}, status=500)


def _reanchor_thumbnails(ws, thumb_dir: str) -> None:
    """Rebuilds every embedded thumbnail from the sidecar files, in current row order.

    openpyxl anchors images to absolute cells, and delete_rows() does NOT move them — after
    a deletion the pictures would stay put while the data shifts up, so every row would show
    the wrong image. Clearing and re-adding from `row_<id>.png` is simpler than trying to
    patch each anchor, and self-heals a sheet whose images already drifted.
    """
    ws._images = []
    for row in ws.iter_rows(min_row=2):
        value = row[COL_ID - 1].value
        if value is None:
            continue
        thumb_path = os.path.join(thumb_dir, f"row_{int(value)}.png")
        if not os.path.isfile(thumb_path):
            continue
        try:
            with Image.open(thumb_path) as im:
                width, height = im.size
            xl_img = XLImage(thumb_path)
            xl_img.width, xl_img.height = width, height
            ws.add_image(xl_img, f"{get_column_letter(COL_THUMB)}{row[0].row}")
        except Exception:
            pass  # a missing/corrupt thumbnail must not block the delete


async def _handle_delete_row(request):
    """Deletes one logged row (and its thumbnail) from a workbook."""
    blocked = _local_only(request)
    if blocked is not None:
        return blocked
    try:
        payload = await request.json()
        path = _resolve_excel_path(payload.get("excel_path", ""))
        row_id = int(payload.get("id", -1))
        if not os.path.isfile(path):
            return web.json_response({"ok": False, "error": "Excel file not found"}, status=404)

        wb = load_workbook(path)
        ws = wb.active
        target = None
        for row in ws.iter_rows(min_row=2):
            if row[COL_ID - 1].value == row_id:
                target = row[0].row
                break
        if target is None:
            return web.json_response({"ok": False, "error": f"Row ID {row_id} not found"}, status=404)

        ws.delete_rows(target)
        thumb_dir = _thumb_dir_for(path)
        _reanchor_thumbnails(ws, thumb_dir)
        _sync_headers(ws)
        _save_with_retry(wb, path)

        # Only remove the sidecar image after the sheet saved — otherwise a failed save
        # would leave the row present but its thumbnail gone.
        try:
            os.remove(os.path.join(thumb_dir, f"row_{row_id}.png"))
        except OSError:
            pass

        _notify_updated(path)
        return web.json_response({"ok": True, "id": row_id})
    except Exception as exc:
        return web.json_response({"ok": False, "error": str(exc)}, status=400)


def _resolve_in_output(subfolder: str, filename: str) -> str:
    """Resolve an image inside ComfyUI's output directory, or raise ValueError.

    Thumbnail replacement is restricted to `output/` on purpose: those are the images this
    workflow actually produced. Same containment rules as the promptDB sandbox — realpath
    first, then commonpath — and the filename is reduced to its basename so it can't carry
    a path of its own.
    """
    base = os.path.realpath(folder_paths.get_output_directory())
    sub = str(subfolder or "").strip().replace("\\", "/")
    if _CONTROL_CHARS.search(sub) or any(p and set(p) == {"."} for p in sub.split("/")):
        raise ValueError("Invalid subfolder")

    name = os.path.basename(str(filename or "").strip())
    if not name or _CONTROL_CHARS.search(name):
        raise ValueError("Invalid filename")
    if os.path.splitext(name)[1].lower() not in {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif"}:
        raise ValueError("Not an image file")

    resolved = os.path.realpath(os.path.join(base, sub, name))
    if not _contain(resolved, base):
        raise ValueError("Thumbnails can only be taken from the output folder")
    if not os.path.isfile(resolved):
        raise ValueError("Image not found")
    return resolved


async def _handle_set_thumbnail(request):
    """Replaces one row's thumbnail with an image picked from output/."""
    blocked = _local_only(request)
    if blocked is not None:
        return blocked
    try:
        payload = await request.json()
        path = _resolve_excel_path(payload.get("excel_path", ""))
        row_id = int(payload.get("id", -1))
        src = _resolve_in_output(payload.get("subfolder", ""), payload.get("filename", ""))
        try:
            size = max(16, min(512, int(payload.get("thumbnail_size", 128))))
        except (TypeError, ValueError):
            size = 128

        if not os.path.isfile(path):
            return web.json_response({"ok": False, "error": "Excel file not found"}, status=404)

        wb = load_workbook(path)
        ws = wb.active
        if not any(row[COL_ID - 1].value == row_id for row in ws.iter_rows(min_row=2)):
            wb.close()
            return web.json_response({"ok": False, "error": f"Row ID {row_id} not found"}, status=404)

        thumb_dir = _thumb_dir_for(path)
        os.makedirs(thumb_dir, exist_ok=True)
        with Image.open(src) as im:
            im = im.convert("RGB")
            im.thumbnail((size, size), Image.Resampling.LANCZOS)
            im.save(os.path.join(thumb_dir, f"row_{row_id}.png"), format="PNG")

        # Rewrite every anchor so the sheet's embedded copy matches the sidecar again.
        _reanchor_thumbnails(ws, thumb_dir)
        _sync_headers(ws)
        _save_with_retry(wb, path)
        _notify_updated(path)
        return web.json_response({
            "ok": True,
            "thumbnail": _thumbnail_data_uri(os.path.join(thumb_dir, f"row_{row_id}.png")),
        })
    except ValueError as exc:
        return web.json_response({"ok": False, "error": str(exc)}, status=400)
    except Exception as exc:
        return web.json_response({"ok": False, "error": str(exc)}, status=400)


async def _handle_download_library(request):
    """Streams a library to the browser as a .zip backup.

    Export means "download a backup", so it is a READ from the sandbox handed to the
    browser — the browser's own download handling decides where it lands. That keeps the
    node free of any write-to-arbitrary-path capability, which is precisely what the
    sandbox exists to prevent.

    The zip carries the workbook plus its thumbnail sidecar folder: the xlsx alone opens
    fine in Excel (openpyxl embeds thumbnails into the sheet), but our gallery reads them
    from the sidecar, so a restore needs both.
    """
    blocked = _local_only(request)
    if blocked is not None:
        return blocked
    try:
        payload = await request.json()
        src = _resolve_excel_path(payload.get("source_path", ""))
        if not os.path.isfile(src):
            return web.json_response({"ok": False, "error": "Library file not found"}, status=404)

        import zipfile

        base = os.path.splitext(os.path.basename(src))[0]
        buf = BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.write(src, os.path.basename(src))
            thumbs = _thumb_dir_for(src)
            if os.path.isdir(thumbs):
                for name in sorted(os.listdir(thumbs)):
                    full = os.path.join(thumbs, name)
                    if os.path.isfile(full):
                        # Mirror the on-disk layout (…/_tj_thumbnails_tmp/<workbook>/row_N.png)
                        # so unzipping into promptDB restores the sidecar where it belongs.
                        zf.write(full, f"{THUMB_SUBDIR}/{base}/{name}")

        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        # Non-ASCII library names would break a bare filename= header, so send an ASCII
        # fallback plus the RFC 5987 UTF-8 form.
        filename = f"{base}_{stamp}.zip"
        ascii_name = re.sub(r"[^A-Za-z0-9._-]", "_", filename)
        return web.Response(
            body=buf.getvalue(),
            headers={
                "Content-Type": "application/zip",
                "Content-Disposition": (
                    f'attachment; filename="{ascii_name}"; '
                    f"filename*=UTF-8''{quote(filename)}"
                ),
            },
        )
    except Exception as exc:
        return web.json_response({"ok": False, "error": str(exc)}, status=400)


if PromptServer is not None:
    PromptServer.instance.routes.post("/tj_node/promptdb/list_rows")(_handle_list_rows)
    PromptServer.instance.routes.post("/tj_node/promptdb/update_row")(_handle_update_row)
    PromptServer.instance.routes.post("/tj_node/promptdb/delete_row")(_handle_delete_row)
    PromptServer.instance.routes.post("/tj_node/promptdb/set_thumbnail")(_handle_set_thumbnail)
    PromptServer.instance.routes.get("/tj_node/promptdb/libraries")(_handle_get_libraries)
    PromptServer.instance.routes.post("/tj_node/promptdb/libraries")(_handle_save_libraries)
    PromptServer.instance.routes.post("/tj_node/promptdb/browse/list")(_handle_browse_list)
    PromptServer.instance.routes.post("/tj_node/promptdb/browse/mkdir")(_handle_browse_mkdir)
    PromptServer.instance.routes.post("/tj_node/promptdb/download_library")(_handle_download_library)
